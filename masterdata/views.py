import os, platform

from django.shortcuts import redirect, render
from django.urls import reverse_lazy, reverse
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, TemplateView
from django.views.generic.edit import UpdateView, DeleteView, FormView
from django.db.models import Q, Sum
from django.http import HttpRequest, HttpResponse, Http404, HttpResponseRedirect, JsonResponse, FileResponse
from django.db import IntegrityError
from django.core.files.storage import FileSystemStorage
from django.utils import timezone

import math

from django_tables2 import LazyPaginator, SingleTableView

import xlsxwriter
import openpyxl
import io

from .models import MasterDataCustomer, MasterDataProduct
from .forms import MasterDataCustomerForm, MasterDataProductForm
from .tables import MasterDataCustomerTable, MasterDataProductTable

# Create your views here.

####################################### START HOME #######################################

class MasterDataHomeView(TemplateView):
    template_name = 'masterdata/md_home.html'

####################################### END HOME #######################################

####################################### START CUSTOMER #######################################

class MasterDataCustomerCreateView(CreateView):
    model = MasterDataCustomer
    form_class = MasterDataCustomerForm
    template_name = 'masterdata/md_customer_input.html'

    def form_valid(self, form):
        address1 = form.cleaned_data['address1']
        address2 = form.cleaned_data['address2']
        temp_form = super(MasterDataCustomerCreateView, self).form_valid(form = form)
        form.instance.address = str(address1) + ', ' + str(address2)
        form.save()
        return temp_form

class MasterDataCustomerListView(SingleTableView):
    model = MasterDataCustomer
    template_name = 'masterdata/md_customer_list.html'
    context_table_name = 'md_customer_table'
    table_class = MasterDataCustomerTable

class MasterDataCustomerUpdateView(UpdateView):
    model = MasterDataCustomer
    form_class = MasterDataCustomerForm
    context_object_name = 'masterdata/md_customer_edit'
    template_name = 'masterdata/md_customer_update.html'

    def get_initial(self):
        initial = super().get_initial()
        address = self.object.address
        address_list = address.split(', ')
        index_split = math.ceil(len(address_list)/2)
        if index_split != 1:
            address1_l, address2_l = address_list[0:index_split], address_list[index_split:]
        else:
            address1_l, address2_l = address_list, []
        address1 = ''
        for i in range(0, len(address1_l)):
            if i == len(address1_l)-1:
                address1 = address1 + str(address1_l[i])
            else:
                address1 = address1 + str(address1_l[i]) + ', '
        address2 = ''
        for i in range(0, len(address2_l)):
            if i == len(address2_l)-1:
                address2 = address2 + str(address2_l[i])
            else:
                address2 = address2 + str(address2_l[i]) + ', '

        initial['address1'] = address1
        initial['address2'] = address2
        return initial

    def form_valid(self, form):
        address1 = form.cleaned_data['address1']
        address2 = form.cleaned_data['address2']
        temp_form = super(MasterDataCustomerUpdateView, self).form_valid(form = form)
        form.instance.address = str(address1) + ', ' + str(address2)
        form.save()
        return temp_form

class MasterDataCustomerDeleteView(DeleteView):
    model = MasterDataCustomer
    template_name = 'masterdata/md_customer_delete.html'
    context_object_name = 'md_customer_delete'
    success_url = reverse_lazy('md_customer_list')

def import_md_customer_xlsx(request):
    if request.method == 'POST':
        request_file = request.FILES.get('import_md_customer_xlsx')
        print('request_file = ', request_file)
        if request_file:
            fs = FileSystemStorage()
            file = fs.save(request_file.name, request_file)
            print('file = ', file)
            wb = openpyxl.load_workbook(filename=file, read_only=True)
            ws = wb.active

            customer_list = []
            for row in ws.iter_rows(min_row=2):
                customer = MasterDataCustomer()
                customer.code = row[0].value
                customer.name = row[1].value
                customer.address = row[2].value
                customer.city = row[3].value

                if customer.code != None and customer.name != None:
                    if customer.address == None:
                        customer.address = '-'
                    elif customer.city == None:
                        customer.city = '-'
                    customer_list.append(customer)
        
            try:
                MasterDataCustomer.objects.bulk_create(customer_list)
                wb.close()
                fs.delete(file)

                url = reverse_lazy('md_customer_list')
                warning = 'Success importing new customers from Excel file.'

                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))
            except IntegrityError as ie:
                wb.close()
                fs.delete(file)
                url = reverse_lazy('md_customer_list')
                print(ie)
                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(ie, url))
    return response

####################################### END CUSTOMER #######################################

####################################### START PRODUCT #######################################

class MasterDataProductCreateView(CreateView):
    form_class = MasterDataProductForm
    template_name = 'masterdata/md_product_input.html'

class MasterDataProductListView(SingleTableView):
    model = MasterDataProduct
    template_name = 'masterdata/md_product_list.html'
    context_table_name = 'md_product_table'
    table_class = MasterDataProductTable

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return MasterDataProduct.objects.filter( Q(part_id_internal__icontains=query) | Q(part_id_customer__icontains=query) | Q(part_name__icontains=query) | Q(material__icontains=query) | Q(customer__name__icontains=query))
        else:
            return MasterDataProduct.objects.all().order_by('customer', 'part_id_internal')

class MasterDataProductUpdateView(UpdateView):
    model = MasterDataProduct
    form_class = MasterDataProductForm
    context_object_name = 'md_product_edit'
    template_name = 'masterdata/md_product_update.html'

class MasterDataProductDeleteView(DeleteView):
    model = MasterDataProduct
    template_name = 'masterdata/md_product_delete.html'
    context_object_name = 'md_product_delete'
    success_url = reverse_lazy('md_product_list')

def import_md_product_xlsx(request):
    if request.method == 'POST':
        request_file = request.FILES.get('import_md_product_xlsx')
        print('request_file = ', request_file)
        if request_file:
            fs = FileSystemStorage()
            file = fs.save(request_file.name, request_file)
            print('file = ', file)
            wb = openpyxl.load_workbook(filename=file, read_only=True)
            ws = wb.active

            product_list = []
            for row in ws.iter_rows(min_row=2):
                product = MasterDataProduct()
                product.part_id_customer = row[0].value
                product.part_id_internal = row[1].value
                customer = row[2].value
                if customer:
                    customer_code = MasterDataCustomer.objects.filter(name = customer).values('code').first()
                    print('customer_code = ', customer_code)
                    if customer_code:
                        product.customer = MasterDataCustomer.objects.get(code = customer_code['code'])
                    else:
                        print('customer_code not on db = ', customer_code)
                        product.customer = MasterDataCustomer.objects.get(code = 'NODATA')
                else:
                    print('customer_id is None from excel = ', customer_code)
                    product.customer = MasterDataCustomer.objects.get(code = 'NODATA')
                product.part_name = row[3].value
                product.material = row[4].value

                if product.part_id_customer != None and product.part_id_internal != None and product.part_name != None and product.material != None:  
                    product_list.append(product)

            try:
                MasterDataProduct.objects.bulk_create(product_list)
                wb.close()
                fs.delete(file)

                url = reverse_lazy('md_product_list')
                warning = 'Success importing new products from Excel file.'

                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))
            except IntegrityError as ie:
                wb.close()
                fs.delete(file)
                url = reverse_lazy('md_product_list')
                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(ie, url))

    return response


####################################### END PRODUCT #######################################
