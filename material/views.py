from django.shortcuts import redirect, render
from django.urls import reverse_lazy, reverse
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, TemplateView
from django.views.generic.list import BaseListView
from django.views.generic.edit import UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.core import signing, serializers
from django.http import HttpRequest, HttpResponse, Http404, HttpResponseRedirect, JsonResponse
from django.core.signing import BadSignature
from django.core.cache import cache
from django.contrib.postgres.search import SearchVector
from django.db.models import Sum
from django.core.files.storage import FileSystemStorage
from django import template, forms
from django.conf import settings

import decimal

from datetime import datetime, date, time, timedelta

import xlsxwriter
import openpyxl
import io

# Custom django filter
register = template.Library()

# Create your views here.
from .models import Product, HopperFillData, Scrap, EstimasiMaterialUsageCO
from .forms import ProductForm, HopperForm, ScrapForm

class MaterialHomeView(TemplateView):
    template_name = 'material/material_home.html'

class ProductListView(ListView):
    model = Product
    template_name = 'material/product.html'
    context_object_name = 'product_list'
    paginate_by = 50

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return Product.objects.filter( Q(part_id__icontains=query) | Q(part_name__icontains=query) | Q(material__icontains=query))
        else:
            return Product.objects.all().order_by('id')
    
    
class ProductDetailView(DetailView):
    model = Product
    template_name = 'material/product_detail.html'

class ProductCreateView(LoginRequiredMixin, CreateView):
    form_class = ProductForm
    template_name = 'material/product_input.html'

class ProductUpdateView(UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'material/product_edit.html'

class ProductDeleteView(DeleteView):
    model = Product
    template_name = 'material/product_delete.html'
    context_object_name = 'product_delete'
    success_url = reverse_lazy('product')

from datetime import datetime, date, time
import pytz

tz = pytz.timezone('Asia/Jakarta')
date_format = '%d/%m/%Y'
time_format = '%H:%M'
shift1_s = time(8, 0, 0)
shift1_e = time(15, 59, 59)
shift2_s = time(16, 0, 0)
shift2_e = time(23, 59, 59)
shift3_s = time(0, 0, 0)
shift3_e = time(7, 59, 59)
shift_choice = ['Shift 1', 'Shift 2', 'Shift 3']

def get_shift(jam_isi):
    if (jam_isi >= shift1_s) and (jam_isi <= shift1_e):
        shift = shift_choice[0]
    elif (jam_isi >= shift2_s) and (jam_isi <= shift2_e):
        shift = shift_choice[1]
    elif (jam_isi >= shift3_s) and (jam_isi <= shift3_e):
        shift = shift_choice[2]
    return shift

class HopperCreateView(LoginRequiredMixin, CreateView):
    form_class = HopperForm
    template_name = 'material/hopper_fill.html'

    def form_valid(self, form):
        no_mesin = form.cleaned_data['no_mesin']
        product = form.cleaned_data['product']
        no_lot = form.cleaned_data['no_lot']
        temp = form.cleaned_data['temp']
        tanggal = form.cleaned_data['tanggal']
        jam_isi = form.cleaned_data['jam_isi']
        co_virgin = form.cleaned_data['co_virgin']
        co_regrind = form.cleaned_data['co_regrind']
        pemakaian_virgin = form.cleaned_data['pemakaian_virgin']
        pemakaian_regrind = form.cleaned_data['pemakaian_regrind']
        pic = form.cleaned_data['pic']
        shift = form.cleaned_data['shift']

        form.instance.shift = get_shift(jam_isi)
        form.instance.pic = self.request.user
        form.save()
        return super(HopperCreateView, self).form_valid(form = form)

def ajax_fill_hopper_form(request):
    no_mesin = request.GET.get('no_mesin', None)
    print(no_mesin)
    if no_mesin:
        try:
            query = HopperFillData.objects.filter(no_mesin=no_mesin).values().latest('id')
            product_id = query['product_id']
            co_virgin = query['co_virgin']
            co_regrind = query['co_regrind']
            pemakaian_virgin = query['pemakaian_virgin']
            pemakaian_regrind = query['pemakaian_regrind']
            if query['kebutuhan_material'] > 0:
                kebutuhan_material = query['kebutuhan_material']
            else:
                kebutuhan_material = None
            temp = query['temp']
            hopper_dict = {'product_id':product_id, 'temp':temp, 'co_virgin':co_virgin, 'co_regrind':co_regrind, 'pemakaian_virgin':pemakaian_virgin, 'pemakaian_regrind':pemakaian_regrind, 'kebutuhan_material': kebutuhan_material}
            return JsonResponse(hopper_dict, safe=False)
        except ValueError:
            pass

class HopperListView(ListView):
    model = HopperFillData
    template_name = 'material/hopper_fill_data.html'
    context_object_name = 'hopper_fill_data_list'
    paginate_by = 50
    ordering = ['-id']

    def get_queryset(self):
        #search_vector = SearchVector('no_mesin', 'product__part_id', 'product__part_name', 'product__material')
        query = self.request.GET.get('qu')
        print(query)
        if (query) and (" " not in query):
            print('query not containing space')
            return HopperFillData.objects.filter( Q(no_mesin__icontains=query) | Q(product__part_id__icontains=query) | Q(product__part_name__icontains=query) | Q(product__material__icontains=query)).order_by('-id')
        elif (query) and (" " in query):
            print('query containing space')
            return HopperFillData.objects.annotate(search = SearchVector('no_mesin', 'product__part_id', 'product__part_name', 'product__material'),).filter(search=query).order_by('-id')
        else:
            return HopperFillData.objects.all().order_by('-id')
    
    def index(request):
        per_page = 50
        page = int(request.GET.get('page', 1))
        start_range = (page - 1) * per_page
        end_range = page * per_page
        records = HopperFillData.objects.all()[start_range:end_range]
        return render(request, 'hopper_fill_data.html', {'records':records})

class HopperUpdateView(UpdateView):
    form_class = HopperForm
    model = HopperFillData
    template_name = 'material/hopper_fill_edit.html'

    def form_valid(self, form):
        no_mesin = form.cleaned_data['no_mesin']
        product = form.cleaned_data['product']
        no_lot = form.cleaned_data['no_lot']
        temp = form.cleaned_data['temp']
        tanggal = form.cleaned_data['tanggal']
        jam_isi = form.cleaned_data['jam_isi']
        co_virgin = form.cleaned_data['co_virgin']
        co_regrind = form.cleaned_data['co_regrind']
        pemakaian_virgin = form.cleaned_data['pemakaian_virgin']
        pemakaian_regrind = form.cleaned_data['pemakaian_regrind']
        pic = form.cleaned_data['pic']
        shift = form.cleaned_data['shift']

        form.instance.shift = get_shift(jam_isi)
        form.instance.pic = self.request.user
        form.save()
        return super(HopperUpdateView, self).form_valid(form = form)

class HopperDeleteView(DeleteView):
    model = HopperFillData
    template_name = 'material/hopper_fill_delete.html'
    context_object_name = 'hopper_delete'
    success_url = reverse_lazy('hopper_fill_data')

class ScrapCreateView(LoginRequiredMixin ,CreateView):
    form_class = ScrapForm
    model = Scrap
    template_name = 'material/scrap_input.html'
    
    def form_valid(self, form):
        tanggal = form.cleaned_data['tanggal']
        shift = form.cleaned_data['shift']
        jumlah_purge = form.cleaned_data['jumlah_purge']
        jumlah_ng = form.cleaned_data['jumlah_ng']
        jumlah_runner = form.cleaned_data['jumlah_runner']
        pic = form.cleaned_data['pic']
        t = super(ScrapCreateView, self).form_valid(form = form)
        form.instance.pic = self.request.user
        form.save()
        return t

class ScrapListView(ListView):
    model = Scrap
    template_name = 'material/scrap_list.html'
    context_object_name = 'scrap_list'
    paginate_by = 50
    ordering = ['-id']

class ScrapUpdateView(UpdateView):
    form_class = ScrapForm
    model = Scrap
    template_name = 'material/scrap_edit.html'

    def form_valid(self, form):
        tanggal = form.cleaned_data['tanggal']
        shift = form.cleaned_data['shift']
        jumlah_purge = form.cleaned_data['jumlah_purge']
        jumlah_ng = form.cleaned_data['jumlah_ng']
        jumlah_runner = form.cleaned_data['jumlah_runner']
        pic = form.cleaned_data['pic']
        t = super(ScrapUpdateView, self).form_valid(form = form)
        form.instance.pic = self.request.user
        form.save()
        return t

class ScrapDeleteView(DeleteView):
    model = Scrap
    template_name = 'material/scrap_delete.html'
    context_object_name = 'scrap_delete'
    success_url = reverse_lazy('scrap_list')

def export_hopper_xlsx(request):
    if request.method == 'POST':
        if request.POST.get('request_date_start_dt'):
            request_date_start = request.POST.get('request_date_start_dt')
            request_date_start = datetime.strptime(request_date_start, '%Y-%m-%d')
        else:
            request_date_start = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')

        if request.POST.get('request_date_end_dt'):
            request_date_end = request.POST.get('request_date_end_dt')
            request_date_end = datetime.strptime(request_date_end, '%Y-%m-%d')
        else:
            request_date_end = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
    else:
        pass

    output = io.BytesIO()
    
    wb = xlsxwriter.Workbook(output, {'in_memory':True})
    ws = wb.add_worksheet(name = 'Material Hopper List')

    row_num = 0

    bold = wb.add_format({'bold':True})
    date_format = wb.add_format({'num_format':'d mmm yyyy'})
    time_format = wb.add_format({'num_format':'hh:mm'})

    columns = [ 'No Mesin', 'Part ID', 'Part Name', 'Material', 'No Lot', 'Temperature (°C)', 'Tanggal', 'Jam Isi',
                 'CO Virgin (kg)', 'CO Regrind (kg)', 'Total CO (kg)', 'Usage Virgin (kg)', 'Usage Regrind (kg)', 'Total Usage (kg)', 'Shift', 'PIC']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], bold)

    filtered_query = HopperFillData.objects.filter(tanggal__gte=request_date_start.date(), tanggal__lte=request_date_end.date())

    if filtered_query:
        no_mesin = filtered_query.values_list('no_mesin', flat=True).order_by('tanggal', 'jam_isi')
        part_id = filtered_query.values_list('product__part_id', flat=True).order_by('tanggal', 'jam_isi')
        part_name = filtered_query.values_list('product__part_name', flat=True).order_by('tanggal', 'jam_isi')
        material = filtered_query.values_list('product__material', flat=True).order_by('tanggal', 'jam_isi')
        no_lot = filtered_query.values_list('no_lot', flat=True).order_by('tanggal', 'jam_isi')
        temp = filtered_query.values_list('temp', flat=True).order_by('tanggal', 'jam_isi')
        tanggal = filtered_query.values_list('tanggal', flat=True).order_by('tanggal', 'jam_isi')
        jam_isi = filtered_query.values_list('jam_isi', flat=True).order_by('tanggal', 'jam_isi')
        co_virgin = filtered_query.values_list('co_virgin', flat=True).order_by('tanggal', 'jam_isi')
        co_regrind = filtered_query.values_list('co_regrind', flat=True).order_by('tanggal', 'jam_isi')
        virgin = filtered_query.values_list('pemakaian_virgin', flat=True).order_by('tanggal', 'jam_isi')
        regrind = filtered_query.values_list('pemakaian_regrind', flat=True).order_by('tanggal', 'jam_isi')
        shift = filtered_query.values_list('shift', flat=True).order_by('tanggal', 'jam_isi')
        pic = filtered_query.values_list('pic', flat=True).order_by('tanggal', 'jam_isi')

        total_co = []
        total_usage = []

        if len(co_virgin) == len(co_regrind):
            for i in range(0, len(co_virgin)):
                co = co_virgin[i] + co_regrind[i]
                total_co.append(co)

        if len(virgin) ==  len(regrind):
            for i in range(0, len(virgin)):
                usage = virgin[i] + regrind[i]
                total_usage.append(usage)

        q_list = [no_mesin, part_id, part_name, material, no_lot, temp, tanggal, jam_isi, co_virgin, co_regrind, total_co, virgin, regrind, total_usage, shift, pic]
        
        for q in range(0, len(q_list)):
            r = 1
            for d in q_list[q]:
                if isinstance(d, str):
                    ws.write_string(r, q, d)
                elif isinstance(d, int):
                    ws.write_number(r, q, d)
                elif isinstance(d, date):
                    ws.write_datetime(r, q, d, date_format)
                elif isinstance(d, time):
                    ws.write_datetime(r, q, d, time_format)
                else:
                    pass
                r += 1

        wb.close()

        output.seek(0)

        response = HttpResponse(output.read(), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename = Material Hopper ' + str(request_date_start.strftime('%d-%m-%Y')) + ' sd ' +  str(request_date_end.strftime('%d-%m-%Y')) + '.xlsx'

        
    else:
        wb.close()
        output.seek(0)

        url = reverse_lazy('hopper_fill_data')
        warning = '0 Entry with date between ' + str(request_date_start.date()) + ' and ' + str(request_date_end.date())

        response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))

    return response

    

def get_material_used(queryset):
    material = queryset.values_list('product__material', flat=True).order_by('tanggal', 'jam_isi')
    tanggal = queryset.values_list('tanggal', flat=True).order_by('tanggal', 'jam_isi')
    virgin = queryset.values_list('pemakaian_virgin', flat=True).order_by('tanggal', 'jam_isi')
    regrind = queryset.values_list('pemakaian_regrind', flat=True).order_by('tanggal', 'jam_isi')

    material_list = []
    for m in material:
        if m not in material_list:
            material_list.append(m)

    #print(material_list)

    material_sum = []
    for m in material_list:
        query = queryset.filter(product__material=m)
        virgin = query.aggregate(Sum('pemakaian_virgin'))
        regrind = query.aggregate(Sum('pemakaian_regrind'))
        material_sum.append([m, virgin, regrind])

    result = []
    for i in range(0, len(material_sum)):
        material = material_sum[i][0]
        virgin = material_sum[i][1]['pemakaian_virgin__sum']
        regrind = material_sum[i][2]['pemakaian_regrind__sum']
        total = virgin + regrind
        result.append([material, virgin, regrind, total])

    #print(result)
    return result
    


def export_material_usage(request):
    if request.method == 'POST':
        if request.POST.get('request_date_start_mu'):
            request_date_start = request.POST.get('request_date_start_mu')
            request_date_start = datetime.strptime(request_date_start, '%Y-%m-%d')
        else:
            request_date_start = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')

        if request.POST.get('request_date_end_mu'):
            request_date_end = request.POST.get('request_date_end_mu')
            request_date_end = datetime.strptime(request_date_end, '%Y-%m-%d')
        else:
            request_date_end = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
    else:
        pass

    output = io.BytesIO()

    start_date = request_date_start.date()
    end_date = request_date_end.date()
    
    wb = xlsxwriter.Workbook(output, {'in_memory':True})
    if start_date == end_date:
        ws = wb.add_worksheet(name = ('Report (' + str(start_date.strftime('%d-%m-%Y'))+')'))
    else:
        ws = wb.add_worksheet(name = ('Report (' + str(start_date.strftime('%d-%m-%Y')) + 'sd' + str(end_date.strftime('%d-%m-%Y'))+')'))

    bold = wb.add_format({'bold':True})
    date_format = wb.add_format({'num_format':'d mmm yyyy'})
    
    filtered_query = HopperFillData.objects.filter(tanggal__gte= start_date,tanggal__lte=end_date)

    daily_material_usage = get_material_used(queryset=filtered_query)

    #untuk ws_1 (daily)
    x, y = 1, 0
    ws.write(0, 0, 'Material', bold)
    ws.write(0, 1, 'Virgin Usage (kg)', bold)
    ws.write(0, 2, 'Regrind Usage (kg)', bold)
    ws.write(0, 3, 'Total Usage (kg)', bold)

    for dl in daily_material_usage:
        ws.write_string(x, y, dl[0])
        ws.write_number(x, y+1, dl[1])
        ws.write_number(x, y+2, dl[2])
        ws.write_number(x, y+3, dl[3])
        x += 1

    wb.close()

    output.seek(0)

    response = HttpResponse(output.read(), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if start_date == end_date:
        response['Content-Disposition'] = 'attachment; filename = Material Usage Report ' + str(start_date.strftime('%d-%m-%Y')) + '.xlsx'
    else:
        response['Content-Disposition'] = 'attachment; filename = Material Usage Report ' + str(start_date.strftime('%d-%m-%Y')) + ' sd ' + str(end_date.strftime('%d-%m-%Y')) + '.xlsx'

    return response

def import_product_xlsx(request):
    if request.method == 'POST':
        request_file = request.FILES.get('import_product_xlsx')
        print('request_file = ', request_file)
        if request_file:
            fs = FileSystemStorage()
            file = fs.save(request_file.name, request_file)
            print('file = ', file)
            wb = openpyxl.load_workbook(filename=file, read_only=True)
            ws = wb.active

            product_list = []
            for row in ws.iter_rows(min_row=2):
                product = Product()
                product.part_id = row[1].value
                product.part_name = row[2].value
                product.material = row[3].value
                product.cycle_time = row[4].value
                product.customer = row[5].value
                product.part_weight = row[6].value
                product.runner_weight = row[7].value
                product.cavity = row[8].value

                if product.part_id != None and product.part_name != None and product.material != None and product.cycle_time != None and product.customer != None and product.part_weight != None and product.runner_weight != None and product.cavity != None:    
                    product_list.append(product)

            Product.objects.bulk_create(product_list)

            wb.close()
            fs.delete(file)

            url = reverse_lazy('product')
            warning = 'Success importing new products from Excel file.'

            response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))
        
    return response



def export_scrap_xlsx(request):
    if request.method == 'POST':
        if request.POST.get('request_date_start'):
            request_date_start = request.POST.get('request_date_start')
            request_date_start = datetime.strptime(request_date_start, '%Y-%m-%d')
        else:
            request_date_start = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')

        if request.POST.get('request_date_end'):
            request_date_end = request.POST.get('request_date_end')
            request_date_end = datetime.strptime(request_date_end, '%Y-%m-%d')
        else:
            request_date_end = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
    else:
        pass

    output = io.BytesIO()
    
    wb = xlsxwriter.Workbook(output, {'in_memory':True})
    ws = wb.add_worksheet(name = 'Scrap')

    row_num = 0

    bold = wb.add_format({'bold':True})
    date_format = wb.add_format({'num_format':'d mmm yyyy'})
    time_format = wb.add_format({'num_format':'hh:mm'})
    decimal_format = wb.add_format({'num_format':'0.00'})

    columns = [ 'Tanggal', 'Shift', 'Purging (kg)', 'Part Scrap (kg)', 'Runner (kg)', 'Total (kg)', 'PIC']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], bold)

    filtered_query = Scrap.objects.filter(tanggal__gte=request_date_start.date(), tanggal__lte=request_date_end.date())

    if filtered_query:
        tanggal = filtered_query.values_list('tanggal', flat=True).order_by('tanggal', 'shift')
        shift = filtered_query.values_list('shift', flat=True).order_by('tanggal', 'shift')
        jumlah_purge = filtered_query.values_list('jumlah_purge', flat=True).order_by('tanggal', 'shift')
        jumlah_ng = filtered_query.values_list('jumlah_ng', flat=True).order_by('tanggal', 'shift')
        jumlah_runner = filtered_query.values_list('jumlah_runner', flat=True).order_by('tanggal', 'shift')
        pic = filtered_query.values_list('pic', flat=True).order_by('tanggal', 'shift')

        total_scrap = []
        if len(jumlah_purge) == len(jumlah_ng) == len(jumlah_runner):
            for i in range(0, len(jumlah_purge)):
                row_scrap = jumlah_purge[i] + jumlah_ng[i] + jumlah_runner[i]
                total_scrap.append(row_scrap)

        q_list = [tanggal, shift, jumlah_purge, jumlah_ng, jumlah_runner, total_scrap, pic]
        
        for q in range(0, len(q_list)):
            r = 1
            for d in q_list[q]:
                if isinstance(d, str):
                    ws.write_string(r, q, d)
                elif isinstance(d, int):
                    ws.write_number(r, q, d)
                elif isinstance(d, float):
                    ws.write_number(r, q, d, decimal_format)
                elif isinstance(d, decimal.Decimal):
                    ws.write_number(r, q, d, decimal_format)
                elif isinstance(d, date):
                    ws.write_datetime(r, q, d, date_format)
                elif isinstance(d, time):
                    ws.write_datetime(r, q, d, time_format)
                else:
                    pass
                r += 1

        wb.close()

        output.seek(0)

        response = HttpResponse(output.read(), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename = Scrap ' + str(request_date_start.strftime('%d-%m-%Y')) + '_-_' +  str(request_date_end.strftime('%d-%m-%Y')) + '.xlsx'

        
    else:
        wb.close()
        output.seek(0)

        url = reverse_lazy('scrap_list')
        warning = '0 Entry with date between ' + str(request_date_start.date()) + ' and ' + str(request_date_end.date())

        response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))

    return response

class MaterialHopperSummary(TemplateView):
    template_name = 'material/hopper_fill_summary.html'

def hopper_line_chart(request):
    if request.method == 'POST':
        if request.POST.get('request_date_start'):
            request_date_start = request.POST.get('request_date_start')
            request_date_start = datetime.strptime(request_date_start, '%Y-%m-%d')
        else:
            request_date_start = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
            request_date_start = request_date_start - timedelta(days=30)
        
        if request.POST.get('request_date_end'):
            request_date_end = request.POST.get('request_date_end')
            request_date_end = datetime.strptime(request_date_end, '%Y-%m-%d')
        else:
            request_date_end = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
        
        date_list = []
        date_end = request_date_end.date()
        date_start = request_date_start.date()
        for d in range((date_end-date_start).days + 1):
            d_between = date_start + timedelta(days=d)
            date_list.append(d_between)
    else:
        lastmonth = datetime.now(tz=tz).date() - timedelta(days=30)
        today = datetime.now(tz=tz).date()
        date_list = []
        for d in range((today-lastmonth).days + 1):
            d_between = lastmonth + timedelta(days=d)
            date_list.append(d_between)

    labels = []
    data_total_virgin = []
    data_total_regrind = []
    data_material_total = []
    data_material_ideal = []

    for dates in date_list:
        labels.append(dates)
        virgin = HopperFillData.objects.filter(tanggal=dates).values('pemakaian_virgin').aggregate(total_virgin=Sum('pemakaian_virgin'))
        total_virgin = virgin['total_virgin']
        data_total_virgin.append(total_virgin)
        regrind = HopperFillData.objects.filter(tanggal=dates).values('pemakaian_regrind').aggregate(total_regrind=Sum('pemakaian_regrind'))
        total_regrind = regrind['total_regrind']
        data_total_regrind.append(total_regrind)
        if total_virgin != None and total_regrind != None:
            material_total = total_virgin+total_regrind
            data_material_total.append(material_total)
        else:
            material_total = 0
            data_material_total.append(material_total)
        berat_ideal = EstimasiMaterialUsageCO.objects.filter(tanggal=dates).values('berat_target_output').aggregate(total_berat=Sum('berat_target_output'))
        total_berat_ideal = berat_ideal['total_berat']
        print(dates, material_total, total_berat_ideal)
        data_material_ideal.append(total_berat_ideal)

    return JsonResponse(data={
        'labels':labels, 'data_total_virgin':data_total_virgin, 'data_total_regrind':data_total_regrind, 'data_material_total':data_material_total, 'data_material_ideal':data_material_ideal,
            })

def hopper_bar_chart(request):
    print(request.method)
    if request.method == 'POST':
        if request.POST.get('request_date_start'):
            request_date_start = request.POST.get('request_date_start')
            print(request_date_start)
            request_date_start = datetime.strptime(request_date_start, '%Y-%m-%d')
        else:
            request_date_start = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
            request_date_start = request_date_start - timedelta(days=30)
        
        if request.POST.get('request_date_end'):
            request_date_end = request.POST.get('request_date_end')
            print(request_date_end)
            request_date_end = datetime.strptime(request_date_end, '%Y-%m-%d')
        else:
            request_date_end = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
        
        date_list = []
        date_end = request_date_end.date()
        date_start = request_date_start.date()
        for d in range((date_end-date_start).days + 1):
            d_between = date_start + timedelta(days=d)
            date_list.append(d_between)
    else:
        date_start = datetime.now(tz=tz).date() - timedelta(days=30)
        date_end = datetime.now(tz=tz).date()
        date_list = []
        for d in range((date_end-date_start).days + 1):
            d_between = date_start + timedelta(days=d)
            date_list.append(d_between)
    
    material = Product.objects.order_by('material').values_list('material', flat=True).distinct()
    material = list(material)
    data_material = []
    data_virgin = []
    data_regrind = []
    for m in material:
        data_material.append(m)
        queryset_v = HopperFillData.objects.filter(product__material=m, tanggal__gte=date_start, tanggal__lte=date_end).values('pemakaian_virgin').aggregate(total_virgin=Sum('pemakaian_virgin'))
        data_virgin.append(queryset_v['total_virgin'])
        queryset_r = HopperFillData.objects.filter(product__material=m, tanggal__gte=date_start, tanggal__lte=date_end).values('pemakaian_regrind').aggregate(total_regrind=Sum('pemakaian_regrind'))
        data_regrind.append(queryset_r['total_regrind'])

    print(len(data_material))

    return JsonResponse(data={
        'data_material':data_material, 'data_virgin':data_virgin, 'data_regrind':data_regrind, 
            })


def hopper_pie_chart(request):
    print(request.method)
    if request.method == 'POST':
        if request.POST.get('request_date_start'):
            request_date_start = request.POST.get('request_date_start')
            print(request_date_start)
            request_date_start = datetime.strptime(request_date_start, '%Y-%m-%d')
        else:
            request_date_start = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
            request_date_start = request_date_start - timedelta(days=30)
        
        if request.POST.get('request_date_end'):
            request_date_end = request.POST.get('request_date_end')
            print(request_date_end)
            request_date_end = datetime.strptime(request_date_end, '%Y-%m-%d')
        else:
            request_date_end = datetime.strptime(datetime.now(tz=tz).strftime('%Y-%m-%d'), '%Y-%m-%d')
        
        date_list = []
        date_end = request_date_end.date()
        date_start = request_date_start.date()
        for d in range((date_end-date_start).days + 1):
            d_between = date_start + timedelta(days=d)
            date_list.append(d_between)
    else:
        date_start = datetime.now(tz=tz).date() - timedelta(days=30)
        date_end = datetime.now(tz=tz).date()
        date_list = []
        for d in range((date_end-date_start).days + 1):
            d_between = date_start + timedelta(days=d)
            date_list.append(d_between)
    

    queryset = HopperFillData.objects.filter(tanggal__gte=date_start, tanggal__lte=date_end).values('pemakaian_virgin').aggregate(total_virgin=Sum('pemakaian_virgin'))
    total_virgin = queryset['total_virgin']
    queryset = HopperFillData.objects.filter(tanggal__gte=date_start, tanggal__lte=date_end).values('pemakaian_regrind').aggregate(total_regrind=Sum('pemakaian_regrind'))
    total_regrind = queryset['total_regrind']
    


    return JsonResponse(data={
                'data_virgin':total_virgin, 'data_regrind':total_regrind, 
            })