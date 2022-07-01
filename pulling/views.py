from django.shortcuts import redirect, render
from django.urls import reverse_lazy, reverse
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, TemplateView
from django.views.generic.list import BaseListView
from django.views.generic.edit import UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db import models
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from django.core import signing, serializers
from django.http import HttpRequest, HttpResponse, Http404, HttpResponseRedirect, JsonResponse, FileResponse
from django.core.signing import BadSignature
from django.core.cache import cache
from django.contrib.postgres.search import SearchVector
from django.db import IntegrityError
from django.core.files.storage import FileSystemStorage
from django import template, forms
from django.conf import settings
from django.utils import timezone

import math
import decimal
from numpy import row_stack
import qrcode

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

from datetime import datetime, date, time, timedelta
import pytz

from django_tables2 import LazyPaginator, SingleTableView

from tablib import Dataset
import xlsxwriter
import openpyxl
import io

# Create your views here.

from .models import PullingCustomer, PullingProduct, PullingFinishGoodItem, FinishGoodStock, TempPullingScanOutModel
from .tables import PullingCustomerTable, PullingFinishGoodItemTable, PullingProductTable, FinishGoodStockTable
from .forms import PullingCustomerForm, PullingProductForm, PullingLabelForm

from .forms import PullingLabelScanInForm, PullingLabelScanOutForm

from .models import TempPullingScanInModel, TempPullingScanOutModel, ScanInModel, ScanOutModel
from .tables import TempScanInTable, TempScanOutTable, ScanInTable, ScanOutTable
from .forms import TempScanInForm, TempScanOutForm

tz = pytz.timezone('Asia/Jakarta')

class PullingHomeView(TemplateView):
    template_name = 'pulling_home.html'

class PullingCustomerCreateView(CreateView):
    model = PullingCustomer
    form_class = PullingCustomerForm
    template_name = 'pulling_customer_input.html'

    def form_valid(self, form):
        address1 = form.cleaned_data['address1']
        address2 = form.cleaned_data['address2']
        temp_form = super(PullingCustomerCreateView, self).form_valid(form = form)
        form.instance.address = str(address1) + ', ' + str(address2)
        form.save()
        return temp_form

class PullingCustomerListView(SingleTableView):
    model = PullingCustomer
    template_name = 'pulling_customer_list.html'
    context_table_name = 'pulling_customer_table'
    table_class = PullingCustomerTable

class PullingCustomerUpdateView(UpdateView):
    model = PullingCustomer
    form_class = PullingCustomerForm
    context_object_name = 'pulling_customer_edit'
    template_name = 'pulling_customer_update.html'

    def get_initial(self):
        initial = super().get_initial()
        address = self.object.address
        address_list = address.split(', ')
        index_split = math.ceil(len(address_list)/2)
        if index_split != 1:
            address1_l, address2_l = address_list[0:index_split], address_list[index_split:]
        else:
            address1_l, address2_l = address_list, []
        address1 = ''
        for i in range(0, len(address1_l)):
            if i == len(address1_l)-1:
                address1 = address1 + str(address1_l[i])
            else:
                address1 = address1 + str(address1_l[i]) + ', '
        address2 = ''
        for i in range(0, len(address2_l)):
            if i == len(address2_l)-1:
                address2 = address2 + str(address2_l[i])
            else:
                address2 = address2 + str(address2_l[i]) + ', '

        initial['address1'] = address1
        initial['address2'] = address2
        return initial

    def form_valid(self, form):
        address1 = form.cleaned_data['address1']
        address2 = form.cleaned_data['address2']
        temp_form = super(PullingCustomerUpdateView, self).form_valid(form = form)
        form.instance.address = str(address1) + ', ' + str(address2)
        form.save()
        return temp_form

class PullingCustomerDeleteView(DeleteView):
    model = PullingCustomer
    template_name = 'pulling_customer_delete.html'
    context_object_name = 'pulling_customer_delete'
    success_url = reverse_lazy('pulling_customer_list')

class PullingProductCreateView(CreateView):
    form_class = PullingProductForm
    template_name = 'pulling_product_input.html'

class PullingProductListView(SingleTableView):
    model = PullingProduct
    template_name = 'pulling_product_list.html'
    context_table_name = 'pulling_product_table'
    table_class = PullingProductTable

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return PullingProduct.objects.filter( Q(part_id_sanwa__icontains=query) | Q(part_id_customer__icontains=query) | Q(part_name__icontains=query) | Q(material__icontains=query) | Q(customer__name__icontains=query))
        else:
            return PullingProduct.objects.all().order_by('part_id_sanwa')

class PullingProductUpdateView(UpdateView):
    model = PullingProduct
    form_class = PullingProductForm
    context_object_name = 'pulling_product_edit'
    template_name = 'pulling_product_update.html'

class PullingProductDeleteView(DeleteView):
    model = PullingProduct
    template_name = 'pulling_product_delete.html'
    context_object_name = 'pulling_product_delete'
    success_url = reverse_lazy('pulling_product_list')

def import_pulling_customer_xlsx(request):
    if request.method == 'POST':
        request_file = request.FILES.get('import_pulling_customer_xlsx')
        print('request_file = ', request_file)
        if request_file:
            fs = FileSystemStorage()
            file = fs.save(request_file.name, request_file)
            print('file = ', file)
            wb = openpyxl.load_workbook(filename=file, read_only=True)
            ws = wb.active

            customer_list = []
            for row in ws.iter_rows(min_row=2):
                customer = PullingCustomer()
                customer.code = row[0].value
                customer.name = row[1].value
                customer.address = row[2].value
                customer.city = row[3].value

                if customer.code != None and customer.name != None:
                    if customer.address == None:
                        customer.address = '-'
                    elif customer.city == None:
                        customer.city = '-'
                    customer_list.append(customer)
        
            try:
                PullingCustomer.objects.bulk_create(customer_list)
                wb.close()
                fs.delete(file)

                url = reverse_lazy('pulling_customer_list')
                warning = 'Success importing new customers from Excel file.'

                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))
            except IntegrityError as ie:
                wb.close()
                fs.delete(file)
                url = reverse_lazy('pulling_customer_list')
                print(ie)
                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(ie, url))
    return response
    
def import_pulling_product_xlsx(request):
    if request.method == 'POST':
        request_file = request.FILES.get('import_pulling_product_xlsx')
        print('request_file = ', request_file)
        if request_file:
            fs = FileSystemStorage()
            file = fs.save(request_file.name, request_file)
            print('file = ', file)
            wb = openpyxl.load_workbook(filename=file, read_only=True)
            ws = wb.active

            product_list = []
            for row in ws.iter_rows(min_row=2):
                product = PullingProduct()
                product.part_id_customer = row[0].value
                product.part_id_sanwa = row[1].value
                customer = row[2].value
                if customer:
                    customer_id = PullingCustomer.objects.filter(name = customer).values('id').first()
                    print('customer_id = ', customer_id)
                    if customer_id:
                        product.customer = PullingCustomer.objects.get(id = customer_id['id'])
                    else:
                        print('customer_id not on db = ', customer_id)
                        product.customer = PullingCustomer.objects.get(id = 1)
                else:
                    print('customer_id is None from excel = ', customer_id)
                    product.customer = PullingCustomer.objects.get(id = 1)
                product.part_name = row[3].value
                product.material = row[4].value
                product.rack = row[5].value
                if type(row[6].value) != int:
                    product.bin_quantity = 0
                else:
                    product.bin_quantity = row[6].value
                if type(row[7].value) != int:
                    product.packaging_quantity = 0
                else:
                    product.packaging_quantity = row[7].value
                if type(row[8].value) != int:
                    product.packaging_quantity = 0
                else:
                    product.full_bin_quantity = row[8].value

                if product.part_id_customer != None and product.part_id_sanwa != None and product.part_name != None and product.material != None:  
                    product_list.append(product)

            try:
                PullingProduct.objects.bulk_create(product_list)
                wb.close()
                fs.delete(file)

                url = reverse_lazy('pulling_product_list')
                warning = 'Success importing new products from Excel file.'

                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))
            except IntegrityError as ie:
                wb.close()
                fs.delete(file)
                url = reverse_lazy('pulling_product_list')
                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(ie, url))

    return response

####################################### Pulling Label ########################################
    
class PullingLabelFormView(FormView):
    form_class = PullingLabelForm
    template_name = 'pulling_label_input.html'

def pulling_label_form_ajax(request):
    part_id_customer = request.GET.get('part_id_customer', None)
    print(part_id_customer)
    if part_id_customer:
        query = PullingProduct.objects.filter(part_id_sanwa=part_id_customer).values().last()
        print(query)
        customer_id = query['customer_id']
        customer = PullingCustomer.objects.filter(id=customer_id).values().first().get('name')
        part_id_sanwa = query['part_id_sanwa']
        part_name = query['part_name']
        rack = query['rack']
        material = query['material']
        full_bin_quantity = query['full_bin_quantity']
        bin_quantity = query['bin_quantity']
        packaging_quantity = query['packaging_quantity']
        pulling_form_dict = {'customer':customer, 'part_id_sanwa':part_id_sanwa, 'part_name':part_name, 
                            'rack':rack, 'material':material, 'full_bin_quantity':full_bin_quantity, 
                            'bin_quantity': bin_quantity, 'packaging_quantity':packaging_quantity,}
        return JsonResponse(pulling_form_dict, safe=False)

def draw_pulling_label(canvas, start_x, start_y, customer, part_name, pn_cust, pn_sanwa, shift, d, m, y, line, rack, qty, material, rohs, remarks, qr_path):
    width, height = A4
    canvas.saveState()
    canvas.translate(0, height)
    canvas.translate(start_x, -start_y)
    #border luar
    canvas.rect(0, 0, 250, -250)
    #kotak atas
    canvas.rect(3, -3, 244, -168)
    
    canvas.drawImage(qr_path, (190+247)/2, -96+21, width=50, preserveAspectRatio=True, anchorAtXY=True)
    
    canvas.line(3, -27, 247, -27)   #garis bawah customer
    canvas.line(3, -51, 247, -51)   #garis bawah partname
    canvas.line(3, -75, 187, -75)   #garis bawah part id cust
    canvas.line(3, -99, 247, -99)   #garis bawah part id sanwa
    canvas.line(3, -123, 247, -123) #garis bawah no lot
    canvas.line(3, -147, 247, -147) #garis bawah rack
    
    canvas.line(78, -27, 78, -123)  #garis kanan tulisan partname
    canvas.line(187,-51, 187, -99)  #garis kiri qr code
    
    #garis pembagi no lot
    canvas.line(111.2, -99, 111.2, -123)
    canvas.line(144.4, -99, 144.4, -123)
    canvas.line(177.6, -99, 177.6, -123)
    canvas.line(210.8, -99, 210.2, -123)

    #garis kanan tulisan rack
    canvas.line(64, -123, 64, -171)
    canvas.line(125, -123, 125, -147)
    canvas.line(186, -123, 186, -147)
    
    #kotak bawah
    canvas.rect(3, -174, 244, -73)
    canvas.line(3, -190, 247, -190)
    canvas.line(84, -174, 84,-247)
    canvas.line(165, -174, 165, -247)
    
    pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
    canvas.setFont('Arial', 10)
    canvas.drawString(5, -41, 'PART NAME :')
    canvas.drawString(5, -65, 'PN CUST. :' )
    canvas.drawString(5, -89, 'PN SANWA :' )
    canvas.drawString(5, -113, 'LOT NO. :' )
    canvas.drawString(5, -140, 'RACK :' )
    canvas.drawString(5, -163, 'MATERIAL :' )
    canvas.drawString(127, -140, 'QTY :' )
    
    pdfmetrics.registerFont((TTFont('ArialBd','ArialBd.ttf')))
    if len(customer) <= 25:
        canvas.setFont('ArialBd', 14)
        canvas.drawCentredString(125, -20, customer)               #tulisan customer
    elif len(customer) <= 30:
        canvas.setFont('ArialBd', 12) 
        canvas.drawCentredString(125, -20, customer)
    else:
        customer_list = customer.split(sep=' ')
        index_split = math.ceil(len(customer_list)/2)
        if index_split != 1:
            c1, c2 = customer_list[0:index_split], customer_list[index_split:]
        else:
            c1, c2 = customer_list, []
        customer1, customer2 = '',''
        for i in range(0, len(c1)):
            customer1 = customer1 + ' ' + c1[i]
        for i in range(0, len(c2)):
            customer2 = customer2 + ' ' + c2[i]
        canvas.setFont('ArialBd', 12)
        canvas.drawCentredString(125, -14, customer1)
        canvas.drawCentredString(125, -25, customer2)
        
    if len(part_name) <= 20:
        canvas.setFont('ArialBd', 14)
        canvas.drawCentredString((250/2)+(78/2), -43, part_name)    #tulisan part_name
    elif len(part_name) <= 30:
        canvas.setFont('ArialBd', 9)
        canvas.drawCentredString((250/2)+(78/2), -43, part_name)
    else:
        part_name_list = part_name.split(sep=' ')
        index_split = math.ceil(len(part_name_list)/2)
        if index_split != 1:
            p1, p2 = part_name_list[0:index_split], part_name_list[index_split:]
        else:
            p1, p2 = part_name_list, []
        part_name1, part_name2 = '',''
        for i in range(0, len(p1)):
            part_name1 = part_name1 + ' ' + p1[i]
        for i in range(0, len(p2)):
            part_name2 = part_name2 + ' ' + p2[i]
        canvas.setFont('ArialBd', 8)
        canvas.drawCentredString((250/2)+(78/2), -37, part_name1)
        canvas.drawCentredString((250/2)+(78/2), -46, part_name2)

    if len(pn_cust)<=15:
        canvas.setFont('Arial', 12)
        canvas.drawCentredString((187/2)+(78/2), -67.5, pn_cust)
    elif len(pn_cust) > 18:
        canvas.setFont('Arial', 8)
        canvas.drawCentredString((187/2)+(78/2), -65, pn_cust)
    else:
        canvas.setFont('Arial', 10)
        canvas.drawCentredString((187/2)+(78/2), -66.25, pn_cust)    #tulisan pn_cust
    canvas.setFont('Arial', 12)
    canvas.drawCentredString((187/2)+(78/2), -91, pn_sanwa)     #tulisan pn_sanwa
    canvas.drawCentredString((187/2), -115, shift)              #tulisan shift / lot col 1
    canvas.drawCentredString(((187/2)+ 33.2), -115, d)          #tulisan day / lot col 2
    canvas.drawCentredString(((187/2)+(2*33.2)), -115, m)       #tulisan month / lot col 3
    canvas.drawCentredString(((187/2)+(3*33.2)), -115, y)       #tulisan year / lot col 4
    canvas.drawCentredString(((187/2)+(4*33.2)), -115, line)    #tulisan line / lot col 5
    
    canvas.setFont('Arial', 9)
    canvas.drawString(67, -140, rack)
    if len(qty) > 10:
        canvas.setFont('Arial', 8)
    else:
        canvas.setFont('Arial', 10)
    canvas.drawString(189, -140, qty)
    if len(material)<27:
        canvas.setFont('Arial', 10)
        canvas.drawString(67, -163, material)
    else:
        material_list = material.split(sep=' ')
        index_split = math.ceil(len(material_list)/2)
        if index_split != 1:
            m1, m2 = material_list[0:index_split], material_list[index_split:]
        else:
            m1, m2 = material_list, []
        material1, material2 = '',''
        for i in range(0, len(m1)):
            material1 = material1 + ' ' + m1[i]
        for i in range(0, len(m2)):
            material2 = material2 + ' ' + m2[i]
        canvas.setFont('Arial', 10)
        canvas.drawString(67, -158, material1)
        canvas.drawString(67, -168, material2)
    
    canvas.drawCentredString((3+84)/2, -185, 'QC APPROVAL')
    canvas.drawCentredString((84+165)/2, -185, 'RoHS')
    canvas.drawCentredString((165+247)/2, -185, 'REMARKS')
    canvas.drawImage(rohs, 125, -218, height=52.5, preserveAspectRatio=True, anchorAtXY=True)
    canvas.drawString(170, -218, remarks)
    canvas.restoreState()

def pulling_label_pdf(request):
    if request.method == 'POST':
        form = PullingLabelForm(request.POST)
        if form.is_valid():
            pdf_buffer = io.BytesIO()
            
            part_id_customer = form.cleaned_data['pulling_product']
            customer = form.cleaned_data['customer']
            date = form.cleaned_data['date']
            part_id_sanwa = form.cleaned_data['part_id_sanwa']
            part_name = form.cleaned_data['part_name']
            tooling = form.cleaned_data['tooling']
            rack = form.cleaned_data['rack']
            material = form.cleaned_data['material']
            full_bin_quantity = form.cleaned_data['full_bin_quantity']
            shift_group = form.cleaned_data['shift_group']
            line = form.cleaned_data['line']
            bin_quantity = form.cleaned_data['bin_quantity']
            packaging_quantity = form.cleaned_data['packaging_quantity']
            remarks = form.cleaned_data['remarks']

            shift_group = shift_group.replace('Shift ','')
            shift_group = shift_group.replace(' Group ', '')

            line = line.replace('Mesin ','')
            line = line.replace(' No. ','')

            date = date.strftime('%d-%m-%y')
            day, month, year = date[0:2], date[3:5], date[6:8]
            no_lot = shift_group + day + month + year + line
            if tooling:
                product_name = part_name + ' ' + tooling
            else:
                product_name = part_name

            if tooling:
                pass
            else:
                tooling = ''
            if rack:
                pass
            else:
                rack = ''
            if remarks:
                pass
            else:
                remarks = ''
            part_info = 'Z1'+str(part_id_customer)+'|'+'Z7'+customer+'|'+'Z2'+no_lot+'|'+'Z3'+str(packaging_quantity)+'|'+'Z4'+material+'|'+'Z5'+rack+' ~ '+tooling+' ~ '+remarks+'|'+'Z6\n'

            qr_path = 'qr.png'
            qr = qrcode.QRCode(version=10, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=4)
            qr.add_data(part_info)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color='black', back_color='white')
            qr_img.save(qr_path)

            rohs_path = 'https://i.ibb.co/zSHbncs/rohs-bw.jpg'

            c = canvas.Canvas(pdf_buffer, A4)

            label_amount = math.ceil(int(bin_quantity) / int(packaging_quantity))
            label_position = [(25, 25), (325, 25), (25, 300), (325, 300), (25, 575), (325, 575)]
            if label_amount <= 6:
                for i in range(0, label_amount):
                    x, y = label_position[i]
                    draw_pulling_label(canvas=c, start_x=x, start_y=y, customer=str(customer), part_name=str(product_name), pn_cust=str(part_id_customer), pn_sanwa=str(part_id_sanwa), shift=str(shift_group), d=day, m=month, y=year, line=line, rack=rack, qty=str(packaging_quantity) + ' PCS', material=str(material), rohs=rohs_path, remarks=str(remarks), qr_path=qr_path)
                c.showPage()
                c.save()
            else:
                page_amount = math.floor(label_amount / 6)
                label_remainder = label_amount - (page_amount * 6)
                for i in range(0, page_amount):
                    for j in range(0, 6):
                        x, y = label_position[j]
                        draw_pulling_label(canvas=c, start_x=x, start_y=y, customer=str(customer), part_name=str(product_name), pn_cust=str(part_id_customer), pn_sanwa=str(part_id_sanwa), shift=str(shift_group), d=day, m=month, y=year, line=line, rack=rack, qty=str(packaging_quantity) + ' PCS', material=str(material), rohs=rohs_path, remarks=str(remarks), qr_path=qr_path)
                    c.showPage()
                if label_remainder > 0:
                    for k in range(0, label_remainder):
                        x, y = label_position[k]
                        draw_pulling_label(canvas=c, start_x=x, start_y=y, customer=str(customer), part_name=str(product_name), pn_cust=str(part_id_customer), pn_sanwa=str(part_id_sanwa), shift=str(shift_group), d=day, m=month, y=year, line=line, rack=rack, qty=str(packaging_quantity) + ' PCS', material=str(material), rohs=rohs_path, remarks=str(remarks), qr_path=qr_path)
                    c.showPage()
                c.save()

            pdf_buffer.seek(0)

            return FileResponse(pdf_buffer, as_attachment=False, filename='label_produk.pdf')

###################################### Pulling Log ############################################

class PullingFinishGoodItemListView(SingleTableView):
    model = PullingFinishGoodItem
    template_name = 'pulling_finish_good_list.html'
    context_table_name = 'pulling_finish_good_table'
    table_class = PullingFinishGoodItemTable
    ordering = ['-date_time']

####################################### Pulling In #############################################

#depracated#
class PullingLabelScanInView(FormView):
    form_class = PullingLabelScanInForm
    template_name = 'pulling_scan_in.html'
#depracated#

def pulling_label_in_decode(request):
    if request.method == 'POST':
        scan_text = request.POST.get('scan_in_text')
        if scan_text:
            try:
                data_list = scan_text.split(sep='|')

                data_clean = []
                for data in data_list:
                    for i in range(1, 8):
                        if i == int(data[1]):
                            data_clean.append(data[2:])

                tfg_item = TempPullingScanInModel()

                if data_clean[1] == '037':
                    data_clean[1] = 'PT. INDONESIA EPSON INDUSTRY'

                tfg_item.part_id_customer = data_clean[0]
                tfg_item.customer = data_clean[1]
                tfg_item.lot_no = data_clean[2]
                if tfg_item.customer != 'PT. INDONESIA EPSON INDUSTRY':
                    tfg_item.line = data_clean[2][8:]
                else:
                    tfg_item.line = ''
                print(tfg_item.line)
                tfg_item.quantity = int(data_clean[3])
                tfg_item.part_name = PullingProduct.objects.filter(part_id_customer=data_clean[0]).values('part_name').last()['part_name']
                tfg_item.material = data_clean[4]
                tfg_item.ref_no = data_clean[5]
                tfg_item.note = data_clean[6]
                tfg_item.date_time =  timezone.now()

                print(tfg_item)

                tfg_item.save()

                response = HttpResponseRedirect(reverse('temp_scan_in'))
                return response
            except:
                e = 'Please re-scan the label.'
                url = reverse_lazy('pulling_label_scan_in')
                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(e, url))
                return response

class TempScanInListView(SingleTableView):
    model = TempPullingScanInModel
    template_name = 'pulling_scan_in.html'
    context_table_name = 'pulling_temp_in_table'
    table_class = TempScanInTable
    ordering = ['-date_time']

class TempScanInUpdateView(UpdateView):
    model = TempPullingScanInModel
    form_class = TempScanInForm
    context_object_name = 'pulling_temp_in_update'
    template_name = 'pulling_temp_in_update.html'
    success_url = reverse_lazy('temp_scan_in')

    def form_valid(self, form):
        line = form.cleaned_data['line']
        part_id_customer = form.cleaned_data['part_id_customer']
        part_name = form.cleaned_data['part_name']
        quantity = form.cleaned_data['quantity']
        temp_form = super(TempScanInUpdateView, self).form_valid(form = form)
        form.instance.line = line
        form.instance.part_id_customer = part_id_customer
        form.instance.part_name = part_name
        form.instance.quantity = quantity
        form.save()
        return temp_form

class TempScanInDeleteView(DeleteView):
    model = TempPullingScanInModel
    template_name = 'pulling_temp_in_delete.html'
    context_object_name = 'pulling_temp_in_delete'
    success_url = reverse_lazy('temp_scan_in')

def scan_in_warehouse_db(request):
    if request.method == 'POST':
        temp_data = TempPullingScanInModel.objects.all().values_list()
        
        for tfg in temp_data:
            print(tfg)
            fg = PullingFinishGoodItem()
            si = ScanInModel()
            fg.part_id_customer = tfg[1]
            si.part_id_customer = tfg[1]
            si.line = tfg[2]
            fg.customer = tfg[3]
            si.customer = tfg[3]
            fg.part_name = tfg[4]
            si.part_name = tfg[4]
            fg.lot_no = tfg[5]
            si.lot_no = tfg[5]
            fg.quantity = int(tfg[6])
            si.quantity = int(tfg[6])
            fg.material = tfg[7]
            si.material = tfg[7]
            fg.ref_no = tfg[8]
            si.ref_no = tfg[8]
            si.rack = PullingProduct.objects.filter(part_id_customer=tfg[1]).values('rack').last()['rack']
            fg.note = tfg[9]
            si.note = tfg[9]
            fg.date_time = tfg[10]
            si.date_time = tfg[10]
            fg.save()
            si.save()

        fgs = FinishGoodStock()
        part_id_customer_list = []
        part_id_cust_q = PullingFinishGoodItem.objects.values_list('part_id_customer', flat=True)
        for p in list(part_id_cust_q):
            if p not in part_id_customer_list:
                part_id_customer_list.append(p)
        for part in part_id_customer_list:
            if not FinishGoodStock.objects.filter(part_id_customer=part).exists():
                fgs.part_id_customer = part
                fgs.part_name = PullingFinishGoodItem.objects.filter(part_id_customer=part).values('part_name').last()['part_name']
                fgs.customer = PullingFinishGoodItem.objects.filter(part_id_customer=part).values('customer').last()['customer']
                fgs.material = PullingFinishGoodItem.objects.filter(part_id_customer=part).values('material').last()['material']
                fgs.total_quantity = int(PullingFinishGoodItem.objects.filter(part_id_customer=part).aggregate(Sum('quantity'))['quantity__sum'])
                fgs.save()
            else:
                update_quantity = int(PullingFinishGoodItem.objects.filter(part_id_customer=part).aggregate(Sum('quantity'))['quantity__sum'])
                FinishGoodStock.objects.filter(part_id_customer=part).update(total_quantity=update_quantity)

        url = reverse_lazy('finish_good_stock')
        warning = 'Success importing scanned pulling label to warehouse database.'

        TempPullingScanInModel.objects.all().delete()

        response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))
    
        return response

class ScanInListView(SingleTableView):
    model = ScanInModel
    template_name = 'pulling_in_list.html'
    table_class = ScanInTable
    context_table_name = 'scan_in_table'
    ordering = ['-date_time']

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return ScanInModel.objects.filter( Q(date_time__icontains=query))
        else:
            return ScanInModel.objects.all().order_by('-date_time')


############################ Pulling Out ##########################################

#depracated#
class PullingLabelScanOutView(FormView):
    form_class = PullingLabelScanOutForm
    template_name = 'pulling_scan_out.html'
#depracated#

def pulling_label_out_decode(request):
    if request.method == 'POST':
        scan_text = request.POST.get('scan_out_text')
        if scan_text:
            try:
                data_list = scan_text.split(sep='|')

                data_clean = []
                for data in data_list:
                    for i in range(1, 8):
                        if i == int(data[1]):
                            data_clean.append(data[2:])

                tfg_item = TempPullingScanOutModel()

                if data_clean[1] == '037':
                    data_clean[1] = 'PT. INDONESIA EPSON INDUSTRY'

                tfg_item.part_id_customer = data_clean[0]
                tfg_item.customer = data_clean[1]
                tfg_item.lot_no = data_clean[2]
                tfg_item.quantity = int(data_clean[3])
                tfg_item.part_name = PullingProduct.objects.filter(part_id_customer=data_clean[0]).values('part_name').last()['part_name']
                tfg_item.material = data_clean[4]
                tfg_item.spq = PullingProduct.objects.filter(part_id_customer=data_clean[0]).values('full_bin_quantity').last()['full_bin_quantity']
                tfg_item.rack = PullingProduct.objects.filter(part_id_customer=data_clean[0]).values('rack').last()['rack']
                tfg_item.note = data_clean[6]
                tfg_item.date_time =  timezone.now()

                print(tfg_item)

                tfg_item.save()

                response = HttpResponseRedirect(reverse('temp_scan_out'))
                return response
            except:
                e = 'Please re-scan the label.'
                url = reverse_lazy('pulling_label_scan_out')
                response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(e, url))
                return response

class TempScanOutListView(SingleTableView):
    model = TempPullingScanOutModel
    template_name = 'pulling_scan_out.html'
    context_table_name = 'pulling_temp_out_table'
    table_class = TempScanOutTable
    ordering = ['-date_time']

#depracated#
class TempScanOutUpdateView(UpdateView):
    model = TempPullingScanOutModel
    form_class = TempScanOutForm
    context_object_name = 'pulling_temp_out_update'
    template_name = 'pulling_temp_out_update.html'
    success_url = reverse_lazy('temp_scan_out')

    def get_initial(self):
        initial = super().get_initial()
        part_id_customer = self.object.part_id_customer
        stock = FinishGoodStock.objects.filter(part_id_customer = part_id_customer).values('total_quantity').last()['total_quantity']
        initial['stock'] = stock
        return initial

    def form_valid(self, form):
        customer = form.cleaned_data['customer']
        part_id_customer = form.cleaned_data['part_id_customer']
        part_name = form.cleaned_data['part_name']
        requirement = form.cleaned_data['requirement']
        quantity = form.cleaned_data['quantity']
        lot_no = form.cleaned_data['lot_no']
        no_of_bin = form.cleaned_data['no_of_bin']
        spq = form.cleaned_data['spq']
        rack = form.cleaned_data['rack']
        temp_form = super(TempScanOutUpdateView, self).form_valid(form = form)
        form.instance.customer = customer
        form.instance.part_id_customer = part_id_customer
        form.instance.part_name = part_name
        form.instance.requirement = requirement
        form.instance.quantity = quantity
        form.instance.lot_no = lot_no
        form.instance.no_of_bin = no_of_bin
        form.instance.spq = spq
        form.instance.rack = rack
        form.save()
        return temp_form
#depracated#

class TempScanOutDeleteView(DeleteView):
    model = TempPullingScanOutModel
    template_name = 'pulling_temp_out_delete.html'
    context_object_name = 'pulling_temp_out_delete'
    success_url = reverse_lazy('temp_scan_out')

def scan_out_warehouse_db(request):
    if request.method == 'POST':
        temp_data = TempPullingScanOutModel.objects.all().values_list()
        
        for tfg in temp_data:
            print(tfg)
            fg = PullingFinishGoodItem()
            so = ScanOutModel()
            fg.part_id_customer = tfg[1]
            so.part_id_customer = tfg[1]
            fg.customer = tfg[2]
            so.customer = tfg[2]
            fg.part_name = tfg[3]
            so.part_name = tfg[3]
            fg.lot_no = tfg[4]
            so.lot_no = tfg[4]
            fg.quantity = -int(tfg[6])
            so.quantity = int(tfg[6])
            fg.material = tfg[7]
            so.material = tfg[7]
            fg.note = tfg[8]
            so.note = tfg[8]
            fg.ref_no = None
            so.spq = tfg[9]
            so.rack = tfg[10]
            fg.date_time = tfg[12]
            so.date_time = tfg[12]
            fg.save()
            so.save()

        fgs = FinishGoodStock()
        part_id_customer_list = []
        part_id_cust_q = PullingFinishGoodItem.objects.values_list('part_id_customer', flat=True)
        for p in list(part_id_cust_q):
            if p not in part_id_customer_list:
                part_id_customer_list.append(p)
        for part in part_id_customer_list:
            if not FinishGoodStock.objects.filter(part_id_customer=part).exists():
                fgs.part_id_customer = part
                fgs.part_name = PullingFinishGoodItem.objects.filter(part_id_customer=part).values('part_name').last()['part_name']
                fgs.customer = PullingFinishGoodItem.objects.filter(part_id_customer=part).values('customer').last()['customer']
                fgs.material = PullingFinishGoodItem.objects.filter(part_id_customer=part).values('material').last()['material']
                fgs.total_quantity = int(PullingFinishGoodItem.objects.filter(part_id_customer=part).aggregate(Sum('quantity'))['quantity__sum'])
                fgs.save()
            else:
                update_quantity = int(PullingFinishGoodItem.objects.filter(part_id_customer=part).aggregate(Sum('quantity'))['quantity__sum'])
                FinishGoodStock.objects.filter(part_id_customer=part).update(total_quantity=update_quantity)

        url = reverse_lazy('finish_good_stock')
        warning = 'Success deducting scanned pulling label from warehouse database.'

        TempPullingScanOutModel.objects.all().delete()

        response = HttpResponse("<script> alert( '%s' ); window.location='%s' </script>" %(warning, url))
    
        return response

class ScanOutListView(SingleTableView):
    model = ScanOutModel
    template_name = 'pulling_out_list.html'
    table_class = ScanOutTable
    context_table_name = 'scan_out_table'
    ordering = ['-date_time']

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return ScanOutModel.objects.filter( Q(date_time__icontains=query))
        else:
            return ScanOutModel.objects.all().order_by('-date_time')


########################################### Finish Good Sum Stock ##########################################

class FinishGoodStockListView(SingleTableView):
    model = FinishGoodStock
    template_name = 'pulling_finish_good_stock.html'
    context_table_name = 'finish_good_stock_table'
    table_class = FinishGoodStockTable

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return FinishGoodStock.objects.filter( Q(part_id_customer__icontains=query) | Q(customer__icontains=query) | Q(material__icontains=query))
        else:
            return FinishGoodStock.objects.all().order_by('customer', 'part_id_customer')



